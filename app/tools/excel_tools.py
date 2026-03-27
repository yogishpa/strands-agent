"""Excel file reading tool."""

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from strands.tools import tool


@tool
def read_excel_sheets(file_path: str) -> dict:
    """
    Read all sheets from an Excel file.

    Opens the file with openpyxl, iterates over every sheet, and reads
    each row into a list of cell values.

    Args:
        file_path: Path to the local .xlsx file.

    Returns:
        {"status": "success", "sheets": {"Sheet1": [[...], ...], ...}}
        or {"status": "error", "error_type": "...", "message": "..."}
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheets = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = []
            for row in worksheet.iter_rows(values_only=True):
                rows.append([cell for cell in row])
            sheets[sheet_name] = rows
        workbook.close()
        return {"status": "success", "sheets": sheets}
    except InvalidFileException:
        return {
            "status": "error",
            "error_type": "InvalidFileError",
            "message": f"The file at {file_path} is not a valid Excel file format.",
        }
    except FileNotFoundError:
        return {
            "status": "error",
            "error_type": "FileNotFoundError",
            "message": f"File not found at {file_path}.",
        }
    except Exception as e:
        return {
            "status": "error",
            "error_type": type(e).__name__,
            "message": f"Unexpected error reading {file_path}: {e}",
        }
